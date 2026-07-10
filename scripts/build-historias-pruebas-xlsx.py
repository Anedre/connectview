#!/usr/bin/env python
"""Genera ARIA-Historias-de-Usuario-y-Casos-de-Prueba.xlsx (openpyxl).
Cobertura EXHAUSTIVA de toda la plataforma ARIA por épicas (E01..E20).
Hojas: Portada · Historias de Usuario · Criterios de Aceptacion · Casos de Prueba
· Matriz de Trazabilidad. Sin fórmulas (catálogo estático) → cero errores.
IDs autonumerados: HU-###, CA-###.n, CP-###.n."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

PLUM = "7A1E52"; MAGENTA = "BC5587"; BAND = "FBF1F5"; INK = "1F2A40"
GREY = "7A879F"; LINE = "E7EAF0"; FONT = "Arial"
thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(c):
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10.5)
    c.fill = PatternFill("solid", fgColor=PLUM)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = BORDER

def cell(c, band=False, center=False):
    c.font = Font(name=FONT, size=10, color=INK)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="top", wrap_text=True)
    if band:
        c.fill = PatternFill("solid", fgColor=BAND)
    c.border = BORDER

def sheet_table(ws, headers, widths, rows, estado_col=None, tall=None):
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
        hdr(ws.cell(row=2, column=j, value=h))
    ws.row_dimensions[2].height = 26
    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            cell(ws.cell(row=3 + i, column=j, value=val), band=(i % 2 == 1), center=(j == 1))
        if tall:
            ws.row_dimensions[3 + i].height = tall
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{2 + len(rows)}"
    if estado_col:
        dv = DataValidation(type="list",
                            formula1='"Pendiente,En progreso,Pasó,Falló,Bloqueado"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(estado_col)
        dv.add(f"{col}3:{col}{2 + len(rows)}")

# ── Contenido: épicas → historias.
# historia = (rol, texto, prioridad, cu_ref, [criterios], [pruebas])
#   criterio = (escenario, dado, cuando, entonces)
#   prueba   = (titulo, tipo, prioridad, precond, pasos, datos, esperado)
def C(e, d, c, t): return (e, d, c, t)
def P(ti, tp, pr, pc, pa, da, ex): return (ti, tp, pr, pc, pa, da, ex)

EPICAS = [
 ("E01","Identidad, cuentas y roles",[
  ("usuario nuevo","registrarme con mi correo y los datos de mi empresa para tener mi propia organización","Must","CU-T-01",
   [C("Registro exitoso","ingreso email, contraseña, nombre y empresa válidos","confirmo el registro","se crea mi organización (tenantId) y quedo como Admin"),
    C("Correo duplicado","uso un correo ya registrado","intento registrarme","el sistema lo rechaza con 'cuenta existente'")],
   [P("Alta de organización nueva","Funcional","Alta","Sin cuenta con ese correo","1. Abrir registro\n2. Datos válidos\n3. Confirmar correo","nuevo@demo.com","tenantId creado; usuario en grupo Admins"),
    P("Registro con correo duplicado","Negativa","Media","Correo ya registrado","1. Registrar con correo existente","correo existente","Mensaje 'cuenta existente'; sin tenant nuevo")]),
  ("usuario","iniciar sesión de forma segura y cerrar sesión cuando termine","Must","CU-T-02",
   [C("Login válido","ingreso credenciales correctas","envío","accedo con la UI de mi rol")],
   [P("Login correcto","Funcional","Alta","Usuario confirmado","1. Ingresar credenciales\n2. Entrar","válidas","Acceso concedido; ID Token con tenantId"),
    P("Login con contraseña incorrecta","Negativa","Media","Usuario existente","1. Ingresar mal la clave","clave errónea","Se rechaza el acceso")]),
  ("administrador","invitar usuarios a mi organización y asignarles un rol","Must","CU-T-01",
   [C("Invitación enviada","ingreso el correo del invitado y su rol","envío la invitación","el invitado recibe el correo y al aceptar entra a MI organización con ese rol")],
   [P("Invitar un agente","Funcional","Alta","Sesión de Admin","1. Invitar correo\n2. Elegir rol Agente\n3. Enviar","agente@demo.com","Invitación enviada; al aceptar queda como Agente del tenant"),
    P("Invitado no ve onboarding de Admin","Negativa","Media","Invitado no-admin","1. Aceptar invitación\n2. Entrar","-","No se le pide conectar Connect; solo lectura de integraciones")]),
  ("administrador","definir permisos por rol (agente, supervisor, admin) para que cada quien vea lo suyo","Must",None,
   [C("Gate por permiso","un agente sin permiso de reportes","abre Reportes","no puede acceder / no ve la opción")],
   [P("Permiso restringe vista","Funcional","Media","Rol Agente","1. Iniciar como Agente\n2. Intentar abrir Admin","-","Acceso denegado a Administración")]),
  ("usuario de una empresa con IdP","entrar con el inicio de sesión corporativo (SSO)","Could","CU-T-02",
   [C("SSO federado","el tenant tiene SSO habilitado","elijo 'Entrar con mi organización'","autentico contra el IdP y accedo sin crear otra clave")],
   [P("Login federado SAML/OIDC","Funcional","Baja","IdP configurado","1. Elegir SSO\n2. Autenticar en el IdP","-","Acceso vía federación; sesión creada")]),
 ]),
 ("E02","Onboarding BYO e integraciones",[
  ("administrador del cliente","conectar mi propia instancia de Amazon Connect sin entregar credenciales","Must","CU-T-03",
   [C("Conexión BYO verificada","apliqué la plantilla y pegué el Role ARN","pulso 'Verificar'","el diagnóstico asume el rol y muestra 'Conectado (N OK · 0 error)'"),
    C("Rol incompleto","el rol carece de permisos","verifico","se listan los checks fallidos con su error")],
   [P("Onboarding BYO feliz","E2E","Alta","Cuenta AWS + instancia Connect","1. Iniciar\n2. Aplicar plantilla\n3. Pegar ARNs\n4. Verificar","Role ARN válido","Config guardada; verificación OK"),
    P("Verificación con permisos faltantes","Negativa","Media","Rol limitado","1. Pegar rol incompleto\n2. Verificar","rol sin permisos","Lista de checks fallidos; no conectado")]),
  ("administrador","conectar Salesforce por OAuth para sincronizar leads y actividad","Should","CU-T-04",
   [C("OAuth Salesforce","completo el consentimiento OAuth","vuelvo a ARIA","se guardan tokens y se descubre el esquema para el mapeo")],
   [P("Conectar Salesforce","Funcional","Media","Org Salesforce","1. Iniciar OAuth\n2. Autorizar","credenciales SF","Estado conectado; mapeo disponible")]),
  ("administrador","registrar uno o varios números de WhatsApp y su cuenta (WABA)","Should","CU-T-04",
   [C("Alta de número","registro número y WABA","guardo","el número queda activo y enrutable por su flujo")],
   [P("Alta de número WhatsApp","Funcional","Media","WABA aprobado","1. Registrar número+WABA\n2. Guardar","+51 9XXXXXXXX","Número activo en vista Ruteo")]),
  ("administrador","conectar mis cuentas de Meta (Facebook/Instagram/Messenger) por auto-servicio","Should",None,
   [C("Conectar con Facebook","autorizo con 'Conectar con Facebook'","selecciono la página/cuenta","IG/Messenger/FB quedan disponibles en la bandeja")],
   [P("Conectar página de Meta","Funcional","Media","Cuenta Meta con página","1. Conectar con Facebook\n2. Elegir cuenta","cuenta Meta","Canal disponible; tokens en Secrets")]),
  ("administrador","conectar Mercado Libre para atender sus mensajes en la bandeja","Could",None,
   [C("Canal ML","autorizo ML por OAuth","llega un mensaje de ML","aparece en la bandeja y puedo responder")],
   [P("Alta de canal Mercado Libre","Funcional","Baja","Cuenta ML","1. Autorizar ML\n2. Recibir mensaje","-","Conversación ML visible y respondible")]),
 ]),
 ("E03","Softphone y telefonía",[
  ("agente","abrir mi softphone para atender llamadas de voz","Must","CU-T-02",
   [C("Softphone listo","tengo sesión y pulso 'Conectar'","completa la federación/login","el softphone queda operativo")],
   [P("Apertura de softphone","Funcional","Alta","Instancia conectada","1. Conectar","-","Softphone disponible; voz operativa")]),
  ("agente","seguir trabajando en canales digitales aunque la voz falle","Must","CU-T-02",
   [C("Degradación sin voz","la conexión de voz falla","abro la app","los canales digitales siguen operativos (softphoneUnavailable)")],
   [P("Voz caída, app operativa","Negativa","Alta","Connect no disponible","1. Forzar fallo de voz","-","Chat/WhatsApp/correo siguen; aviso claro")]),
  ("agente","evitar el softphone duplicado cuando abro ARIA en varias pestañas","Should",None,
   [C("Guard multipestaña","ya tengo el softphone en otra pestaña","abro una 2ª","la 2ª ofrece 'Usar aquí' y no duplica el softphone")],
   [P("Guard multipestaña","Funcional","Media","Softphone abierto","1. Abrir 2ª pestaña","-","Ofrece 'Usar aquí'; single softphone")]),
  ("agente","cambiar mi estado (Disponible/No disponible/ACW) para controlar el ruteo","Should",None,
   [C("Cambio de estado","estoy Disponible","me pongo No disponible","dejo de recibir contactos nuevos")],
   [P("Cambio de estado","Funcional","Media","Softphone activo","1. Cambiar a No disponible","-","No entran contactos nuevos")]),
 ]),
 ("E04","Escritorio del agente (omnicanal)",[
  ("agente","recibir el contacto con el contexto 360° del cliente para no buscar en otras pantallas","Must","CU-T-06",
   [C("Contexto al aceptar","acepto un contacto","se abre el workspace","veo perfil, historial y transcripción en vivo")],
   [P("Atención con 360°","Funcional","Alta","Contacto en cola","1. Aceptar\n2. Revisar paneles","contacto de prueba","Cargan perfil/historial/transcripción sin cambiar de pantalla")]),
  ("agente","transferir o poner en conferencia a otro agente o cola","Must",None,
   [C("Transferencia","estoy en un contacto","transfiero a una cola","el contacto pasa con su contexto")],
   [P("Transferir contacto","Funcional","Media","Contacto activo","1. Transferir a cola\n2. Confirmar","-","Contacto ruteado; contexto preservado")]),
  ("agente","recibir sugerencias del copiloto durante la conversación","Should","CU-T-10",
   [C("Copiloto en vivo","estoy en la conversación","pido sugerencias","el copiloto propone réplicas y próxima acción")],
   [P("Sugerencias del copiloto","Funcional","Media","Conversación activa","1. Abrir copiloto","-","Réplicas y próxima acción sugeridas")]),
  ("agente","crear un lead o una tarea desde el contacto sin salir del workspace","Should",None,
   [C("Crear desde contacto","estoy atendiendo","creo un lead/tarea","queda ligado al cliente y a la conversación")],
   [P("Crear lead desde contacto","Funcional","Baja","Contacto activo","1. Crear lead","datos mínimos","Lead creado y vinculado")]),
  ("agente","cerrar con tipificación y notas asistidas por IA","Should","CU-T-10",
   [C("Wrap-up asistido","termino el contacto","acepto la tipificación sugerida","se guarda el wrap-up y se actualiza el perfil")],
   [P("Cierre con tipificación","Funcional","Media","Contacto en curso","1. Finalizar\n2. Aceptar tipificación","-","Wrap-up guardado; perfil actualizado")]),
 ]),
 ("E05","Inbox omnicanal y conversaciones",[
  ("agente","ver todas las conversaciones (WhatsApp, IG, Messenger, FB, ML) en una sola bandeja","Must","CU-T-07",
   [C("Bandeja unificada","hay conversaciones de varios canales","abro la bandeja","las veo todas con su canal y estado")],
   [P("Bandeja omnicanal","Funcional","Alta","Conversaciones en varios canales","1. Abrir bandeja","-","Se listan todas con su canal")]),
  ("agente","responder cada conversación por su propio canal desde la bandeja","Must",None,
   [C("Respuesta por canal","abro una conversación de WhatsApp","respondo","el mensaje sale por WhatsApp al cliente")],
   [P("Responder en la bandeja","Funcional","Alta","Conversación abierta","1. Escribir\n2. Enviar","texto","Mensaje entregado por el canal correcto")]),
  ("agente","ver el Cliente 360 de la conversación (quién es, su historial)","Should","CU-T-06",
   [C("Cliente 360","abro una conversación","reviso el panel de cliente","veo su identidad e historial consolidado")],
   [P("Panel Cliente 360","Funcional","Media","Conversación con cliente conocido","1. Abrir panel cliente","-","Identidad e historial visibles")]),
  ("sistema","cerrar conversaciones de forma consistente (agente, Agente IA o inactividad)","Should","CU-T-13",
   [C("Cierre unificado","el agente cierra o pasan ~10 min inactiva","corre el cierre","la conversación queda cerrada y con métricas")],
   [P("Cierre por reaper","Funcional","Media","Conversación inactiva","1. Dejar inactiva >10 min","-","Reaper la cierra; no queda colgada")]),
 ]),
 ("E06","Supervisión en vivo",[
  ("supervisor","monitorear la cola y el estado de los agentes en tiempo real","Must",None,
   [C("Monitor en vivo","abro Cola en vivo","observo","veo contactos en cola, espera y estado de cada agente")],
   [P("Tablero de cola en vivo","Funcional","Alta","Operación activa","1. Abrir Cola en vivo","-","Métricas y estados en tiempo real")]),
  ("supervisor","escuchar, susurrar o entrar a una llamada para acompañar al agente","Should",None,
   [C("Whisper/Barge","un agente está en llamada","elijo susurrar","hablo con el agente sin que el cliente me oiga")],
   [P("Susurro a agente","Funcional","Media","Llamada en curso","1. Monitor\n2. Susurrar","-","El agente me oye; el cliente no")]),
  ("supervisor","ver la salud de mis agentes (carga, tiempos) para balancear","Could",None,
   [C("Salud de agentes","abro el tablero de agentes","reviso","veo carga y tiempos por agente")],
   [P("Salud de agentes","Funcional","Baja","Datos disponibles","1. Abrir salud de agentes","-","Carga/tiempos por agente")]),
 ]),
 ("E07","Grabaciones e historial",[
  ("supervisor","reabrir una conversación pasada con su grabación y transcripción","Must",None,
   [C("Reproducción","abro una grabación","reproduzco","escucho el audio y leo la transcripción sincronizada")],
   [P("Reproducir grabación","Funcional","Alta","Grabación disponible","1. Abrir grabación\n2. Reproducir","-","Audio + transcripción visibles")]),
  ("supervisor","ver el sentimiento a lo largo de la conversación (heatmap)","Could",None,
   [C("Heatmap de sentimiento","abro una grabación con análisis","reviso la línea de tiempo","veo el sentimiento por tramo")],
   [P("Heatmap por sentimiento","Funcional","Baja","Contact Lens disponible","1. Abrir grabación","-","Línea de tiempo por sentimiento")]),
  ("supervisor","buscar rápidamente una grabación por cliente, agente o contenido","Should",None,
   [C("Búsqueda","abro la búsqueda (⌘K)","escribo un término","obtengo las grabaciones que coinciden")],
   [P("Búsqueda de grabaciones","Funcional","Media","Grabaciones indexadas","1. ⌘K\n2. Buscar","nombre de cliente","Resultados relevantes")]),
  ("supervisor","descargar el audio de una grabación","Could",None,
   [C("Descarga","abro una grabación","descargo el audio","obtengo el archivo (con Content-Disposition correcto)")],
   [P("Descargar audio","Funcional","Baja","Grabación en S3 con CORS/headers","1. Descargar","-","Archivo descargado sin error")]),
 ]),
 ("E08","Leads, embudo y Programas",[
  ("vendedor","ver mi embudo de leads por etapas","Must",None,
   [C("Embudo","abro Leads","observo","veo los leads por etapa con su información")],
   [P("Ver embudo","Funcional","Alta","Leads cargados","1. Abrir Leads","-","Leads por etapa")]),
  ("vendedor","mover un lead de etapa y ver su historial de golpes","Must","CU-T-06",
   [C("Ledger de golpes","abro un lead","reviso su historial","veo cada gestión/WA/correo/llamada como un golpe con resultado")],
   [P("Historial de golpes","Funcional","Media","Lead con actividad","1. Abrir lead\n2. Ver historial","lead activo","Golpes cronológicos con resultado")]),
  ("administrador","clasificar los leads por Programa (unidad de negocio) automáticamente","Must",None,
   [C("Auto-tag de programa","llega un lead de un origen ligado a un programa","se procesa","queda etiquetado en ese Programa")],
   [P("Auto-clasificación por programa","Funcional","Alta","Programa configurado","1. Ingresar lead del origen","lead de prueba","Lead etiquetado en el Programa")]),
  ("vendedor","capturar un lead manualmente en segundos","Should","CU-T-05",
   [C("Captura rápida","abro QuickCapture","ingreso teléfono y datos","el lead queda creado y clasificado")],
   [P("Captura manual (QuickCapture)","Funcional","Media","Sesión activa","1. QuickCapture\n2. Guardar","teléfono+nombre","Lead creado")]),
  ("supervisor","priorizar leads por un puntaje para atacar primero los más calientes","Could",None,
   [C("Scoring","abro el embudo","ordeno por puntaje","los leads con mayor score aparecen primero")],
   [P("Ordenar por scoring","Funcional","Baja","Scoring habilitado","1. Ordenar por score","-","Orden por puntaje descendente")]),
 ]),
 ("E09","Ingesta y speed-to-lead",[
  ("responsable de marketing","que los leads de Meta entren solos y se contacten al instante","Must","CU-T-05",
   [C("Ingesta + speed-to-lead","llega un leadgen de Meta","se procesa","se crea el lead y se dispara el primer contacto en segundos"),
    C("Deduplicación","el teléfono ya existe","llega el leadgen","se fusiona el historial, no se duplica")],
   [P("Lead nuevo end-to-end","E2E","Alta","Página Meta + formulario","1. Enviar leadgen de prueba","lead de prueba","Lead creado, clasificado, WA de bienvenida + tarea"),
    P("Lead duplicado","Negativa","Media","Teléfono existente","1. Enviar leadgen repetido","teléfono repetido","No duplica; fusiona historial")]),
  ("administrador","importar una base de leads por CSV","Should",None,
   [C("Importación CSV","subo un CSV con leads","confirmo","los leads se crean y clasifican")],
   [P("Importar CSV","Funcional","Media","CSV válido","1. Subir CSV\n2. Mapear\n3. Importar","csv de prueba","Leads creados; errores reportados")]),
 ]),
 ("E10","Campañas salientes y dialer",[
  ("supervisor","lanzar una campaña de voz con modo y reglas para contactar sin saturar","Must","CU-T-09",
   [C("Campaña de voz","configuré audiencia, modo y reglas","activo","el marcador respeta pacing/% abandono y reparte a agentes")],
   [P("Campaña de voz predictiva","E2E","Alta","Audiencia + agentes","1. Crear campaña voz\n2. Activar","100 contactos","Marca con pacing; KPIs en vivo")]),
  ("supervisor","lanzar una campaña por WhatsApp con plantilla aprobada","Must","CU-T-09",
   [C("Campaña WhatsApp","tengo una plantilla aprobada","activo la campaña","se envía desde el número del tenant con botones")],
   [P("Campaña WhatsApp","Funcional","Alta","Plantilla aprobada","1. Crear campaña WA\n2. Activar","plantilla","Envíos desde el número del tenant")]),
  ("supervisor","correr varias campañas a la vez con prioridad y metas","Should","CU-T-09",
   [C("Orquestación","hay varias campañas activas","corre el dialer","reparte capacidad por prioridad, peso y metas")],
   [P("Blend multi-campaña","Funcional","Media","2+ campañas activas","1. Activar varias\n2. Observar","-","Reparto por prioridad/peso")]),
  ("supervisor","pausar y reanudar una campaña en cualquier momento","Should",None,
   [C("Control","una campaña está corriendo","la pauso","deja de marcar hasta que la reanude")],
   [P("Pausar/reanudar campaña","Funcional","Media","Campaña activa","1. Pausar\n2. Reanudar","-","Se detiene y retoma correctamente")]),
  ("supervisor","ver los KPIs de la campaña en vivo (contactados, conversión, AHT)","Should",None,
   [C("KPIs en vivo","la campaña corre","abro sus estadísticas","veo contactados, conversión y AHT al momento")],
   [P("KPIs de campaña","Funcional","Media","Campaña con actividad","1. Abrir stats","-","KPIs actualizados")]),
 ]),
 ("E11","Bots (constructor y ejecución)",[
  ("administrador","construir un bot arrastrando pasos en un lienzo, sin programar","Must",None,
   [C("Constructor visual","abro el editor de bots","agrego y conecto pasos","el flujo se arma y se guarda")],
   [P("Crear un flujo de bot","Funcional","Alta","Sesión de Admin","1. Abrir editor\n2. Agregar pasos\n3. Conectar\n4. Guardar","-","Flujo válido guardado")]),
  ("administrador","ramificar el flujo (botones, condición, A/B) según la respuesta","Should",None,
   [C("Ramas","un paso tiene varias salidas","conecto cada salida","el bot enruta por la rama correcta")],
   [P("Ramas por botones","Funcional","Media","Nodo con botones","1. Conectar cada salida","-","Ruteo por rama correcto")]),
  ("administrador","probar el bot antes de publicarlo","Should",None,
   [C("Prueba","abro 'Probar'","envío mensajes","el bot responde como en producción")],
   [P("Probar bot","Funcional","Media","Bot en borrador","1. Probar\n2. Conversar","mensajes","Respuestas esperadas")]),
  ("cliente por WhatsApp","que el bot me atienda solo con la info oficial","Must","CU-T-07",
   [C("Bot resuelve","escribo una consulta cubierta","el bot responde","da la respuesta correcta desde el número del cliente")],
   [P("Consulta resuelta por bot","Funcional","Alta","Bot publicado","1. Enviar consulta conocida","'¿cuánto cuesta?'","Respuesta correcta")]),
 ]),
 ("E12","Agente IA (RAG y herramientas)",[
  ("cliente","que el Agente IA responda en lenguaje natural citando la fuente","Must","CU-T-07",
   [C("RAG con citación","pregunto algo cubierto por la base","el agente responde","da la respuesta citando la fuente ([F]/[C]/[P])"),
    C("Fallback","pregunto algo fuera de la base","el agente no halla match","responde con fallback controlado o deriva")],
   [P("Respuesta con citación","Funcional","Alta","Base cargada","1. Preguntar algo conocido","pregunta","Respuesta correcta con citación"),
    P("Consulta sin cobertura","Negativa","Media","Base cargada","1. Preguntar fuera de alcance","pregunta rara","Fallback; sin inventar")]),
  ("cliente","que el Agente IA ejecute acciones (agendar, consultar) con herramientas","Should",None,
   [C("Herramienta","pido agendar una cita","el agente usa la herramienta","la cita queda registrada")],
   [P("Agente IA usa herramienta","Funcional","Media","Herramienta habilitada","1. Pedir agendar","fecha","Cita creada por el agente IA")]),
  ("cliente","hablar con una persona sin repetir lo que ya conté","Must","CU-T-08",
   [C("Handoff con contexto","pido un asesor","el agente IA deriva","el humano recibe el contacto con el historial")],
   [P("Derivación a humano","Funcional","Alta","Agentes disponibles","1. Pedir asesor","-","Handoff con contexto")]),
 ]),
 ("E13","Automatizaciones y Journeys",[
  ("administrador","definir reglas 'cuando pase X, haz Y' para automatizar tareas","Should","CU-T-11",
   [C("Regla ejecutada","hay una regla activa y ocurre su evento","el motor evalúa","ejecuta las acciones resolviendo tokens [[name]]/{{name}}"),
    C("Sin coincidencia","ninguna regla coincide","ocurre el evento","no se ejecuta ninguna acción (no-op)")],
   [P("Trigger→acción","Funcional","Media","Regla 'nuevo lead→notificar+plantilla'","1. Crear lead\n2. Verificar","lead de prueba","Notifica y envía plantilla con tokens resueltos"),
    P("Evento sin regla","Negativa","Baja","Sin reglas aplicables","1. Disparar evento no cubierto","-","No-op")]),
  ("administrador","armar recorridos multi-paso (mensaje→espera→condición) que avancen solos","Could","CU-T-12",
   [C("Journey avanza","un contacto está inscrito y vence la espera","corre el runner","pasa al siguiente paso"),
    C("Salida temprana","el contacto convierte o pide baja","corre el runner","abandona el journey")],
   [P("Avance de journey","Funcional","Media","Journey publicado","1. Inscribir\n2. Esperar tick","contacto","Avanza de paso"),
    P("Salida por conversión","Funcional","Baja","Contacto en journey","1. Marcar conversión","-","Sale del journey")]),
 ]),
 ("E14","Correo",[
  ("agente","enviar un correo al cliente desde ARIA con el proveedor configurado","Should",None,
   [C("Envío multi-proveedor","el tenant configuró SES/SMTP/Gmail/Microsoft","envío un correo","se entrega por el proveedor del tenant")],
   [P("Enviar correo","Funcional","Media","Proveedor configurado","1. Redactar\n2. Enviar","destinatario de prueba","Correo entregado; registrado en el hilo")]),
  ("agente","responder dentro del hilo de correo del cliente","Could",None,
   [C("Respuesta en hilo","abro un correo del cliente","respondo","la respuesta queda en el mismo hilo")],
   [P("Responder hilo","Funcional","Baja","Hilo existente","1. Responder","texto","Respuesta en el hilo")]),
  ("administrador","usar plantillas de correo con variables","Could",None,
   [C("Plantilla","elijo una plantilla","la envío","las variables se reemplazan con los datos del cliente")],
   [P("Enviar con plantilla","Funcional","Baja","Plantilla creada","1. Elegir plantilla\n2. Enviar","lead con nombre","Variables resueltas")]),
 ]),
 ("E15","WhatsApp (gestión)",[
  ("administrador","crear y gestionar plantillas HSM de WhatsApp","Should",None,
   [C("CRUD de plantillas","creo una plantilla","la envío a aprobación","queda registrada con su estado")],
   [P("Crear plantilla HSM","Funcional","Media","WABA conectado","1. Crear plantilla\n2. Enviar a aprobación","plantilla","Plantilla creada; estado visible")]),
  ("administrador","agregar botones a las plantillas (URL dinámica, copiar, flow, OTP)","Could",None,
   [C("Botones","agrego un botón de URL dinámica","guardo","la plantilla incluye el botón con su parámetro")],
   [P("Plantilla con botones","Funcional","Baja","Plantilla en edición","1. Agregar botón URL\n2. Guardar","url {{1}}","Botón presente y válido")]),
  ("administrador","enrutar cada número de WhatsApp a un flujo distinto","Should",None,
   [C("Ruteo por número","tengo varios números","asigno un flujo a cada uno","cada número entra a su flujo")],
   [P("Ruteo por número","Funcional","Media","2+ números","1. Asignar flujos\n2. Enviar a cada número","-","Cada número usa su flujo")]),
  ("administrador","operar WhatsApp tanto en modo AWS como Meta directo","Could",None,
   [C("Dual-mode","el número es Meta-standalone","envío una plantilla","se envía por Graph API con la cuenta correcta")],
   [P("Envío dual-mode Meta","Funcional","Baja","Número Meta directo","1. Enviar plantilla","cuenta Meta","Enviado por Graph API")]),
 ]),
 ("E16","Citas, callbacks y tareas",[
  ("agente","agendar una cita o un callback con el cliente","Should",None,
   [C("Agendar","estoy con el cliente","agendo una cita","el cliente recibe recordatorio y yo la veo en mi agenda")],
   [P("Agendar cita","Funcional","Media","Contacto/lead activo","1. Agendar\n2. Confirmar","fecha/hora","Cita creada; recordatorio programado")]),
  ("agente","gestionar mis tareas y seguimientos (vencidas, hoy, próximas)","Should",None,
   [C("Tareas","abro mis tareas","reviso","las veo agrupadas por vencidas/hoy/próximas")],
   [P("Lista de tareas","Funcional","Media","Tareas asignadas","1. Abrir Tareas","-","Agrupadas por tiempo")]),
  ("agente","recibir el recordatorio de una cita a tiempo","Could",None,
   [C("Recordatorio","tengo una cita próxima","llega la hora del recordatorio","recibo el aviso")],
   [P("Recordatorio de cita","Funcional","Baja","Cita agendada","1. Esperar recordatorio","-","Aviso recibido")]),
 ]),
 ("E17","Reportes y analítica",[
  ("supervisor","ver tableros por dominio con comparación contra el período previo","Should","CU-T-15",
   [C("KPIs con delta","abro Reportes","cargan los tableros","veo KPIs comparados con el período anterior (DeltaChip)")],
   [P("Tableros con delta","Funcional","Media","Actividad registrada","1. Abrir Reportes","rango","KPIs con comparación de período")]),
  ("supervisor","segmentar los reportes por Programa/campaña","Should",None,
   [C("Por programa","filtro por un programa","aplico","los KPIs corresponden solo a ese programa")],
   [P("Reporte por programa","Funcional","Media","Datos por programa","1. Filtrar por programa","programa","KPIs segmentados")]),
  ("supervisor","descargar reportes en Excel/CSV","Should","CU-T-15",
   [C("Exportación","elijo un reporte","descargo","obtengo el archivo correcto")],
   [P("Descargar reporte","Funcional","Baja","Datos disponibles","1. Elegir\n2. Descargar","-","Excel/CSV correcto")]),
  ("analista","conectar un feed seguro a Power BI","Could",None,
   [C("Feed BI","tengo un token de feed","consulto un dataset","recibo el JSON del dataset sin exponer credenciales")],
   [P("Feed Power BI","Funcional","Baja","Token generado","1. GET con token+dataset","token válido","JSON del dataset; HMAC OK")]),
  ("supervisor","medir el rendimiento de los bots (resolución, derivaciones)","Could",None,
   [C("Reporte de bot","abro el reporte de bots","reviso","veo resolución, derivaciones y volúmenes")],
   [P("Reporte de bot","Funcional","Baja","Conversaciones de bot","1. Abrir get-bot-report","-","Métricas de bot")]),
 ]),
 ("E18","Cumplimiento y deliverability",[
  ("oficial de cumplimiento","que ningún número en 'no molestar' sea contactado","Must","CU-T-14",
   [C("Bloqueo por supresión","un número está suprimido","se intenta un contacto saliente","se bloquea y se registra el motivo"),
    C("Fail-open","el servicio de supresión no responde","se intenta un contacto","no bloquea pero avisa (no frena la operación)")],
   [P("Contacto a número suprimido","Negativa","Alta","Número en DNC","1. Intentar WA/voz saliente","número DNC","Bloqueado; motivo registrado"),
    P("Supresión no disponible","Negativa","Media","Servicio caído","1. Intentar contacto","-","No bloquea; deja aviso")]),
  ("oficial de cumplimiento","no contactar a quien ya convirtió o pidió baja en Salesforce","Should",None,
   [C("Opt-out","el lead tiene opt-out en Salesforce","se intenta contactarlo","queda excluido")],
   [P("Respeta opt-out de SF","Negativa","Media","Lead con opt-out","1. Intentar contacto","-","Excluido del contacto")]),
  ("administrador","ver el estado de salud/cuarentena de mis números de WhatsApp","Could",None,
   [C("Salud WhatsApp","abro el estado de números","reviso","veo calidad/estado y cuarentenas")],
   [P("Estado de salud WA","Funcional","Baja","Números conectados","1. Abrir estado WA","-","Calidad/estado por número")]),
 ]),
 ("E19","IA transversal",[
  ("agente","obtener un resumen y tipificación automáticos al terminar la llamada","Should","CU-T-10",
   [C("Resumen","termina la llamada con transcripción","abro wrap-up","la IA muestra resumen, tipificación y próxima acción"),
    C("Degradación","Bedrock no responde","pido el resumen","mensaje suave y puedo redactar manual")],
   [P("Resumen de llamada","Funcional","Media","Contacto con transcripción","1. Finalizar\n2. Abrir wrap-up","llamada","Resumen+tipificación+próxima acción"),
    P("Fallo de Bedrock","Negativa","Baja","Bedrock caído","1. Pedir resumen","-","Mensaje suave; sin bloqueo")]),
  ("usuario","usar el copiloto global para navegar y ejecutar acciones con IA","Could",None,
   [C("Copiloto global","abro el copiloto","hago una pregunta","responde con bloques interactivos (ir/ejecutar/tip/kpi)")],
   [P("Copiloto responde rico","Funcional","Baja","Permiso use_copilot","1. Preguntar al copiloto","pregunta","Respuesta con botones de navegar/ejecutar")]),
 ]),
 ("E20","Administración y configuración",[
  ("administrador","configurar la plataforma (marca, canales, parámetros) desde un solo lugar","Should",None,
   [C("Configuración","abro Configuración","cambio un parámetro","se guarda y aplica")],
   [P("Guardar configuración","Funcional","Media","Sesión de Admin (Bearer)","1. Cambiar parámetro\n2. Guardar","-","Cambio persistido")]),
  ("administrador","estimar y ver mi consumo/costos (Connect, Meta, plataforma)","Could",None,
   [C("Consumo","abro la Calculadora de Consumo","reviso","veo estimado vs real por concepto")],
   [P("Calculadora de consumo","Funcional","Baja","Datos de uso","1. Abrir Consumo","-","Estimado vs real")]),
  ("usuario","recibir notificaciones (campana) y marcarlas como vistas","Could",None,
   [C("Notificaciones","tengo notificaciones nuevas","abro la campana","las veo y puedo marcarlas como vistas")],
   [P("Campana de notificaciones","Funcional","Baja","Eventos generados","1. Abrir campana\n2. Marcar visto","-","Contador se actualiza")]),
  ("administrador","gestionar usuarios y sus roles/permisos","Should",None,
   [C("Gestión de usuarios","abro Administración","cambio el rol de un usuario","el permiso se aplica")],
   [P("Cambiar rol de usuario","Funcional","Media","Usuarios existentes","1. Cambiar rol\n2. Guardar","-","Rol/permiso aplicado")]),
 ]),
]

PRIOR = {"Must": 0, "Should": 1, "Could": 2, "Won't": 3}

# ── Aplanar con IDs autonumerados ──
HU, CA, CP, TZ = [], [], [], []
n = 0
for ep, epname, stories in EPICAS:
    for rol, texto, prio, cu, crits, tests in stories:
        n += 1
        hu = f"HU-{n:03d}"
        HU.append([hu, ep, epname, f"Como {rol}, quiero {texto}.", prio, cu or "—"])
        cas = []
        for k, (esc, dado, cuando, ent) in enumerate(crits, 1):
            ca = f"CA-{n:03d}.{k}"; cas.append(ca)
            CA.append([ca, hu, esc, f"Dado que {dado}", f"Cuando {cuando}", f"Entonces {ent}"])
        for k, (ti, tp, tpr, pc, pa, da, ex) in enumerate(tests, 1):
            cp = f"CP-{n:03d}.{k}"
            CP.append([cp, ti, ep, hu, cu or "—", tp, tpr, pc, pa, da, ex, "Pendiente"])
            TZ.append([epname, cu or "—", hu, ", ".join(cas), cp, ti])

wb = Workbook()

# Portada
ws = wb.active; ws.title = "Portada"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 112
def prow(r, t, s, c, b=True, h=16):
    cc = ws.cell(row=r, column=2, value=t)
    cc.font = Font(name=FONT, size=s, bold=b, color=c)
    cc.alignment = Alignment(wrap_text=True, vertical="center"); ws.row_dimensions[r].height = h
prow(3, "ARIA · by Novasys", 12, MAGENTA, h=20)
prow(4, "Historias de Usuario, Criterios de Aceptación y Casos de Prueba", 19, PLUM, h=32)
prow(5, "Cobertura integral de la plataforma ARIA por épicas (E01–E20). Actualizado 2026-07-10.", 11, INK, b=False, h=20)
prow(7, "Contenido de este libro", 13, PLUM, h=22)
idx = [
 ("Historias de Usuario", f"{len(HU)} historias en 20 épicas — «Como… quiero… para…» con prioridad MoSCoW."),
 ("Criterios de Aceptación", f"{len(CA)} escenarios — formato Dado / Cuando / Entonces (Gherkin)."),
 ("Casos de Prueba", f"{len(CP)} casos — pasos, datos y resultado esperado; Estado editable."),
 ("Matriz de Trazabilidad", "Épica → Caso de uso → Historia → Criterios → Caso de prueba."),
]
r = 8
for name, desc in idx:
    a = ws.cell(row=r, column=2, value=f"•  {name}"); a.font = Font(name=FONT, size=11, bold=True, color=INK); r += 1
    d = ws.cell(row=r, column=2, value=f"     {desc}"); d.font = Font(name=FONT, size=10, color=GREY)
    d.alignment = Alignment(wrap_text=True); r += 1
prow(r+1, "Épicas: E01 Identidad · E02 Onboarding/Integraciones · E03 Softphone · E04 Escritorio del agente · E05 Inbox omnicanal · E06 Supervisión · E07 Grabaciones · E08 Leads/Programas · E09 Ingesta · E10 Campañas/Dialer · E11 Bots · E12 Agente IA · E13 Automatizaciones/Journeys · E14 Correo · E15 WhatsApp · E16 Citas/Tareas · E17 Reportes · E18 Cumplimiento · E19 IA transversal · E20 Administración.", 9.5, GREY, b=False, h=54)
prow(r+2, "Prioridad (MoSCoW): Must · Should · Could.   Tipo de prueba: Funcional · Negativa · E2E.   Estado: Pendiente · En progreso · Pasó · Falló · Bloqueado.", 9.5, GREY, b=False, h=16)

ws = wb.create_sheet("Historias de Usuario")
HU.sort(key=lambda x: (x[1], PRIOR.get(x[4], 9), x[0]))
sheet_table(ws, ["ID","Épica","Nombre de épica","Historia de usuario","Prioridad","Caso de uso"],
            [10, 8, 22, 74, 11, 12], HU)

ws = wb.create_sheet("Criterios de Aceptacion")
sheet_table(ws, ["ID","Historia","Escenario","Dado","Cuando","Entonces"],
            [12, 10, 22, 34, 28, 40], CA)

ws = wb.create_sheet("Casos de Prueba")
sheet_table(ws, ["ID","Título","Épica","Historia","Caso de uso","Tipo","Prioridad",
                 "Precondición","Pasos","Datos","Resultado esperado","Estado"],
            [11, 26, 8, 10, 12, 10, 10, 24, 32, 15, 36, 12], CP, estado_col=12, tall=52)

ws = wb.create_sheet("Matriz de Trazabilidad")
sheet_table(ws, ["Épica","Caso de uso","Historia","Criterios","Caso de prueba","Título del caso de prueba"],
            [22, 12, 10, 20, 12, 40], TZ)

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "docs", "tecnico", "ARIA-Historias-de-Usuario-y-Casos-de-Prueba.xlsx")
wb.save(OUT)
print("OK", OUT)
print(f"Epicas=20 Historias={len(HU)} Criterios={len(CA)} CasosDePrueba={len(CP)} Traza={len(TZ)}")
